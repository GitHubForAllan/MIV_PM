# -*- coding: utf-8 -*-
"""把兩份現行的採購 Excel 轉成採購追蹤系統的種子資料 purchase_seed.json。

來源：
  1. 品保設備採購進度追蹤_最終版.xlsx
     取「採購進度管制表」的案件與步驟、「預計付款」的金額幣別與各期款。
  2. 订单情况- TÌNH TRẠNG ĐƠN HÀNG.xlsx
     取「跟踪订货契约单」的訂單與付款、「厂商」的中越對照廠商主檔。

用法（預設路徑為越南廠現行檔案位置）：
    python tools/gen_purchase_seed.py
    python tools/gen_purchase_seed.py --qa <品保表.xlsx> --po <訂單表.xlsx> -o purchase_seed.json

需求：pip install openpyxl

產出的 JSON 由 purchase-admin.html 的「Excel 資料匯入」頁選檔載入。
注意：檔案含廠商、金額與付款條件，已在 firebase.json 排除，不會發布到網站上。

轉檔時的判斷規則（與 purchase-flow.js 一致）：
  - 路徑：已取得訂貨契約單號者為 B 流程；其餘依「預付款 < 30 萬台幣
    且 總價 < 100 萬台幣」判斷，缺金額則留空（匯入時預設為 A）。
  - 步驟：品保表依「目前步驟」欄推導；訂單表僅在有單號時推到 B3／B4，
    其餘步驟一律留給人工確認。
  - 幣別：訂單表原本沒有幣別欄，金額達百萬級視為 VND，其餘視為 USD，
    因此這批一律標記 needsReview=true。
  - 採購單位：訂單表的「請購單位」寫法有二十幾種，原文保留在 unitRaw，
    僅語意明確者放進 unitMapSuggested 當匯入畫面的預設值。
"""
import argparse
import datetime
import json
import re

import openpyxl

DEFAULT_QA = r"D:/My Data/allan.yang/Desktop/品保設備採購進度追蹤_最終版.xlsx"
DEFAULT_PO = r"P:/ENG/A02.采购项目管制-QUẢN LÝ MUA HÀNG/订单情况- TÌNH TRẠNG ĐƠN HÀNG.xlsx"

parser = argparse.ArgumentParser(description="產生採購追蹤系統的種子資料")
parser.add_argument("--qa", default=DEFAULT_QA, help="品保設備採購進度追蹤 xlsx 路徑")
parser.add_argument("--po", default=DEFAULT_PO, help="订单情况 TÌNH TRẠNG ĐƠN HÀNG xlsx 路徑")
parser.add_argument("-o", "--out", default="purchase_seed.json", help="輸出的 JSON 檔名")
args = parser.parse_args()

QA = args.qa
PO = args.po

A_STEPS = ["1", "2", "A3", "A4", "A5", "A6", "A7", "A8", "A9", "A10", "A11"]
B_STEPS = ["1", "2", "B1", "B2", "B3", "B4"]


def s(v):
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.date):
        return v.isoformat()
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def iso(v):
    """盡量解析成 YYYY-MM-DD；解析不出來回傳 (None, 原字串)。"""
    if v is None or s(v) == "":
        return None, ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return s(v), ""
    t = s(v)
    m = re.match(r"^(\d{4})[./-](\d{1,2})[./-](\d{1,2})", t)
    if m:
        y, mo, d = map(int, m.groups())
        try:
            return datetime.date(y, mo, d).isoformat(), ""
        except ValueError:
            return None, t
    m = re.match(r"^(\d{1,2})[./-](\d{1,2})$", t)   # 8/27 視為 2026 年
    if m:
        mo, d = map(int, m.groups())
        try:
            return datetime.date(2026, mo, d).isoformat(), ""
        except ValueError:
            return None, t
    return None, t          # 「7月下旬」「6月下旬」之類保留原文


def num(v):
    if v is None or s(v) == "":
        return None
    try:
        return float(str(v).replace(",", "").replace(" ", ""))
    except ValueError:
        return None


def truthy(v):
    return s(v).lower().strip(" .*") not in ("", "-")


def split_bi(t):
    """把「中文 / 換行 / Tieng Viet」拆成兩段。"""
    parts = [p.strip() for p in s(t).split("\n") if p.strip()]
    if len(parts) >= 2:
        return parts[0], " ".join(parts[1:])
    return (parts[0] if parts else ""), ""


# 與 purchase-flow.js 的 suggestPath 相同的門檻規則：
# 預付款 < 30 萬台幣 且 總價 < 100 萬台幣 -> B 訂貨契約單，否則 A 一般合約
FX_TO_TWD = {"TWD": 1.0, "USD": 32.0, "VND": 0.00125}


def suggest_path(amount, currency, prepay_ratio):
    rate = FX_TO_TWD.get(currency)
    if not rate or not amount or amount <= 0:
        return ""
    total_twd = amount * rate
    prepay_twd = total_twd * (prepay_ratio if prepay_ratio and prepay_ratio > 0 else 1)
    return "B" if (prepay_twd < 300000 and total_twd < 1000000) else "A"


# 1. 廠商主檔（來自「厂商」工作表的中越對照）
wb_po = openpyxl.load_workbook(PO, data_only=True)
vendors = []
seen = set()
for row in wb_po["厂商"].iter_rows(min_row=1, max_col=1):
    t = s(row[0].value)
    if not t:
        continue
    if "-" in t:
        zh, vi = t.split("-", 1)
    else:
        zh, vi = t, ""
    zh, vi = zh.strip(), vi.strip()
    if zh and zh not in seen:
        seen.add(zh)
        vendors.append({"name": zh, "nameVi": vi})

# 2. 品保設備追蹤表 -> 完整案件
wb_qa = openpyxl.load_workbook(QA, data_only=True)
ws = wb_qa["採購進度管制表"]
cases = []
for r in range(6, ws.max_row + 1):
    vendor = s(ws.cell(r, 3).value)
    item = s(ws.cell(r, 4).value)
    if not vendor and not item:
        continue
    path = s(ws.cell(r, 5).value).upper() or "A"
    cur = num(ws.cell(r, 6).value)
    keys = A_STEPS if path == "A" else B_STEPS
    steps = {}
    if cur:
        n = int(cur)
        for i, k in enumerate(keys, start=1):
            if i < n:
                steps[k] = {"state": "done"}
            elif i == n:
                steps[k] = {"state": "doing"}
    req, req_raw = iso(ws.cell(r, 19).value)
    eta, eta_raw = iso(ws.cell(r, 20).value)
    remark = s(ws.cell(r, 23).value)
    calc = s(ws.cell(r, 24).value)          # 例 10657.7*40%=4263.08
    unit = "品保課"
    if "總務" in item or "總務" in vendor:
        unit = "總務課"
    elif "製造課" in item or "技術課" in item:
        unit = "生技課"
    notes = [x for x in [
        remark,
        ("金額計算：" + calc) if calc else "",
        ("必要日期原文：" + req_raw) if req_raw else "",
        ("預計完成日原文：" + eta_raw) if eta_raw else "",
    ] if x]
    cases.append({
        "source": "品保設備採購進度追蹤",
        "unit": unit, "vendor": vendor, "item": item, "itemVi": "",
        "path": path, "steps": steps,
        "requiredDate": req, "etaDate": eta,
        "currency": "", "amount": None, "payments": [],
        "poNo": "", "remark": "\n".join(notes),
        "status": "open", "imported": True, "needsReview": False,
    })

# 3. 預計付款表 -> 補金額 幣別 各期款
pay_by_key = {}
wsp = wb_qa["預計付款"]
for r in range(4, wsp.max_row + 1):
    vendor = s(wsp.cell(r, 1).value)
    if not vendor:
        continue
    unit_raw = s(wsp.cell(r, 2).value)
    equip_zh, _vi = split_bi(wsp.cell(r, 4).value)
    usd, vnd = num(wsp.cell(r, 5).value), num(wsp.cell(r, 6).value)
    currency = "USD" if usd else ("VND" if vnd else "")
    amount = usd or vnd
    payments = []
    for i, c in enumerate([7, 9, 11, 13], start=1):     # I. II. III. VI. 的 % 與期限
        pct, term = num(wsp.cell(r, c).value), s(wsp.cell(r, c + 1).value)
        if pct is None and not term:
            continue
        payments.append({
            "seq": i, "pct": pct, "term": term,
            "amount": round(amount * pct, 2) if (amount and pct) else None,
            "dueDate": None, "paidDate": None, "note": "",
        })
    rec = {"currency": currency, "amount": amount, "payments": payments,
           "poNo": s(wsp.cell(r, 3).value), "unitRaw": unit_raw}
    pay_by_key.setdefault(vendor, rec)
    if equip_zh:
        pay_by_key.setdefault(equip_zh, rec)

UNIT_MAP = {"品保": "品保課", "制造": "生技課", "製造": "生技課",
            "实验室": "品保課", "實驗室": "品保課",
            "技術課": "生技課", "技术课": "生技課"}

for c in cases:
    for key in (c["vendor"], c["item"]):
        rec = pay_by_key.get(key)
        if rec:
            c["currency"] = c["currency"] or rec["currency"]
            c["amount"] = c["amount"] or rec["amount"]
            c["payments"] = rec["payments"]
            c["poNo"] = c["poNo"] or rec["poNo"]
            if rec["unitRaw"] in UNIT_MAP:
                c["unit"] = UNIT_MAP[rec["unitRaw"]]
            break

# 4. 跟踪订货契约单 -> 待覆核匯入清單
ws2 = wb_po["跟踪订货契约单"]
orders = []
for r in range(5, ws2.max_row + 1):
    vendor = s(ws2.cell(r, 2).value)
    order_txt = s(ws2.cell(r, 5).value)
    if not vendor and not order_txt:
        continue
    code_raw = s(ws2.cell(r, 3).value)
    m = re.search(r"\((TW|VN)請購\)", code_raw)
    po_no = re.sub(r"\s*\((TW|VN)請購\)\s*", "", code_raw).strip()
    origin = m.group(1) if m else ""
    # 編碼欄偶爾被填成「请购单」之類的說明文字，只有長得像單號（含 6 位以上
    # 連續數字，例如 2026051902/RX-TQ）才算真的取到訂貨契約單號
    if po_no and not re.search(r"\d{6,}", po_no):
        po_note, po_no = po_no, ""
    else:
        po_note = ""
    item_zh, item_vi = split_bi(order_txt)
    amount = num(ws2.cell(r, 6).value)
    # 原表無幣別欄位，金額達百萬級視為 VND，其餘視為 USD，匯入前須人工覆核
    currency = "VND" if (amount and amount >= 1000000) else ("USD" if amount else "")
    payments, pay_raw = [], []
    for i, c in enumerate([10, 14, 18], start=1):
        pct, term = num(ws2.cell(r, c).value), s(ws2.cell(r, c + 1).value)
        amt = num(ws2.cell(r, c + 2).value)
        paid, paid_raw = iso(ws2.cell(r, c + 3).value)
        if pct is None and not term and amt is None and not paid and not paid_raw:
            continue
        if paid_raw:
            pay_raw.append("第%d期付款日原文：%s" % (i, paid_raw))
        payments.append({"seq": i, "pct": pct, "term": term, "amount": amt,
                         "dueDate": None, "paidDate": paid, "note": ""})
    done_order = truthy(ws2.cell(r, 7).value)
    # 有訂貨契約單編號即代表 B3 已取號；訂單已完成則 B4 用印亦完成。其餘步驟留給人工確認。
    steps = {}
    if po_no:
        for k in ["1", "2", "B1", "B2", "B3"]:
            steps[k] = {"state": "done"}
        steps["B4"] = {"state": "done" if done_order else "doing"}
    eta, eta_raw = iso(ws2.cell(r, 25).value)
    notes = [x for x in [
        s(ws2.cell(r, 9).value),
        ("月結付款：" + s(ws2.cell(r, 22).value)) if s(ws2.cell(r, 22).value) else "",
        ("付款條件變更：" + s(ws2.cell(r, 23).value)) if s(ws2.cell(r, 23).value) else "",
        ("付款憑證單：" + s(ws2.cell(r, 24).value)) if s(ws2.cell(r, 24).value) else "",
        ("預計/送貨時間原文：" + eta_raw) if eta_raw else "",
        ("關務資料：" + s(ws2.cell(r, 26).value)) if s(ws2.cell(r, 26).value) else "",
        ("請購來源：" + origin) if origin else "",
        ("編碼欄原文：" + po_note) if po_note else "",
    ] + pay_raw if x]
    orders.append({
        "source": "跟踪订货契约单",
        "unit": "", "unitRaw": s(ws2.cell(r, 4).value),   # 原文保留，匯入時由人工對應到正式單位
        "vendor": vendor, "item": item_zh, "itemVi": item_vi,
        # 已取得訂貨契約單號者必為 B 流程；其餘依金額套用同一門檻規則判斷
        "path": "B" if po_no else suggest_path(amount, currency, payments[0]["pct"] if payments else None),
        "steps": steps,
        "currency": currency, "amount": amount,
        "poNo": po_no, "payments": payments,
        "requiredDate": None, "etaDate": eta,
        "remark": "\n".join(notes),
        "status": "open", "imported": True, "needsReview": True,
    })

# 原表「請購單位」的實際寫法統計，供匯入畫面做對應用
unit_raw_counts = {}
for o in orders:
    k = o["unitRaw"].strip()
    if k:
        unit_raw_counts[k] = unit_raw_counts.get(k, 0) + 1
unit_raw_counts = dict(sorted(unit_raw_counts.items(), key=lambda kv: -kv[1]))

# 原表「請購單位」寫法 -> 正式採購單位的建議對應。
# 只列出語意明確的；製造相關、倉庫、開發與人名等留空，由管理員在匯入畫面自行指定。
UNIT_RAW_MAP = {
    "品保": "品保課",
    "工具室": "工具室",
    "技術課": "生技課",
    "技术课": "生技課",
    "技术课 - 阿倫": "生技課",
    "技术课 -xiaochen": "生技課",
}
unit_map_suggested = {k: UNIT_RAW_MAP.get(k, "") for k in unit_raw_counts}

seed = {
    "generatedAt": datetime.date.today().isoformat(),
    "note": "由「品保設備採購進度追蹤_最終版.xlsx」與「订单情况- TÌNH TRẠNG ĐƠN HÀNG.xlsx」轉出；needsReview=true 者請於匯入前人工覆核採購單位與幣別。",
    "units": [
        {"name": "品保課", "nameVi": "Bộ phận QA", "order": 1},
        {"name": "工具室", "nameVi": "Phòng dụng cụ", "order": 2},
        {"name": "生技課", "nameVi": "Bộ phận Kỹ thuật SX", "order": 3},
        {"name": "總務課", "nameVi": "Bộ phận Hành chính", "order": 4},
    ],
    "vendors": vendors,
    "cases": cases,
    "orders": orders,
    "unitRawCounts": unit_raw_counts,
    "unitMapSuggested": unit_map_suggested,
}
with open(args.out, "w", encoding="utf-8") as f:
    json.dump(seed, f, ensure_ascii=False, indent=1)
print("已寫出", args.out)
print("vendors:", len(vendors), "| cases:", len(cases), "| orders:", len(orders))
print("cases with amount:", sum(1 for c in cases if c["amount"]))
print("orders with poNo:", sum(1 for o in orders if o["poNo"]),
      "| with amount:", sum(1 for o in orders if o["amount"]))
print("unitRaw counts:", unit_raw_counts)
mapped = {}
for o in orders:
    u = UNIT_RAW_MAP.get(o["unitRaw"].strip(), "")
    if u:
        mapped[u] = mapped.get(u, 0) + 1
print("建議對應後各單位筆數:", mapped)
print("尚未對應筆數:", sum(1 for o in orders if not UNIT_RAW_MAP.get(o["unitRaw"].strip(), "")))
